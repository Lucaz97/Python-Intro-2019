import asyncio
from pyppeteer import launch
from openpyxl import load_workbook


async def main(urls):
    prices = []
    # inizialize background browser
    browser = await launch()
    # initialize new page
    page = await browser.newPage()
    # set page size
    await page.setViewport({"width": 1920, "height": 1080})
    for url in urls:
        print("getting", url)
        try:
            # load url
            await page.goto(url)
        except:
            print("Invalid url, skipping...")
            prices.append("Invalid url")
            continue
        try:
            # get price
            price = await page.evaluate('document.getElementsByClassName("_j1kt73")[2].textContent')
            print("Price:", price)
            prices.append(price)
        except:
            # sometimes page don't load properly, try one more time
            await page.goto(url)
            try:
                price = await page.evaluate('document.getElementsByClassName("_j1kt73")[2].textContent')
                print("Price:", price)
                prices.append(price)
            except:
                print("No price found for this url")
                prices.append("No price found")

    await browser.close()
    return prices


def get_urls(file_name, col_urls):
    print("Reading urls from", file_name)
    # carico file
    wb = load_workbook(file_name)
    # carico foglio di lavoro attivo
    ws = wb.active
    # recupero colonna desiderata
    col = ws[col_urls]
    # costruisco e ritorno una lista con gli url
    urls = [cell.value for cell in col if cell.value is not None]
    print("Found", len(urls),"urls.")
    wb.close()
    return urls


def save_data(xlsx_file, prices, col_price):
    print("Saving data on", xlsx_file)
    # carico file
    wb = load_workbook(xlsx_file)
    # carico foglio di lavoro attivo
    ws = wb.active
    # ciclo sulla lista di coppie da inserire
    for i, price in enumerate(prices):  # ['a','b','c'] --enumerate--> [(0,'a'),(1,'b'),(2,'c')]
        # inserisco dati nel file excel, gli indici delle righe su excel partono da uno
        ws[col_price+str(i+1)] = price

    # sovrascrivo il file
    wb.save(xlsx_file)

urls = get_urls('sample.xlsx', "A")
prices = asyncio.get_event_loop().run_until_complete(main(urls))
save_data('sample.xlsx', prices, "B")



